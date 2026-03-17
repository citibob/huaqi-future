#!/usr/bin/env python3
"""
单卡最终决策报告生成器

目标：
- 主清单去杂，但不漏数据
- 保留全量结果与杂质回收
- 输出电商决策需要的核心字段
"""

import json
import re
from collections import defaultdict
from datetime import datetime
from statistics import median

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


JPY_TO_CNY = 0.04322
TARGET_RARITIES = {"PSA10", "BGS10", "ARS10", "AR", "SAR"}
NOISE_TERMS = (
    "まとめ", "まとめ売り", "セット", "引退", "連番", "&", "＆", " psa9", "psa 9",
    "bgs9", "ars9", "grade 9", "9枚", "2枚", "3枚", "4枚", "5枚", "コンプ", "ファイル",
    "バインダー", "スリーブ", "コイン", "サプライ", "box", "ボックス", "パック", "未開封"
)
SECTION_ALIASES = {
    "イーブイ": ["イーブイ", "eevee"],
    "エーフィ": ["エーフィ", "espeon"],
    "ブラッキー": ["ブラッキー", "umbreon", "月亮伊布"],
    "リーフィア": ["リーフィア", "leafeon"],
    "グレイシア": ["グレイシア", "glaceon"],
    "ニンフィア": ["ニンフィア", "sylveon", "仙子伊布"],
    "ゲンガー": ["ゲンガー", "gengar", "耿鬼"],
    "リザードン": ["リザードン", "charizard", "喷火龙"],
    "カメックス": ["カメックス", "blastoise", "水箭龟"],
    "フシギダネ": ["フシギダネ", "bulbasaur", "妙蛙种子"],
    "ルギア": ["ルギア", "lugia"],
    "レックウザ": ["レックウザ", "rayquaza"],
    "サーナイト": ["サーナイト", "gardevoir"],
    "ミュウツー": ["ミュウツー", "mewtwo", "超梦"],
    "ミュウ": ["ミュウ", "mew", "梦幻"],
    "ピカチュウ": ["ピカチュウ", "pikachu", "皮卡丘"],
    "オーガポン": ["オーガポン", "ogerpon"],
    "多龙梅西亚": ["多龙", "ドラメ", "ドラメシヤ", "dreepy"],
    "耿鬼ex": ["耿鬼", "ゲンガー", "gengar"],
    "火伊布": ["ブースター", "flareon", "火伊布"],
    "水伊布": ["シャワーズ", "vaporeon", "水伊布"],
    "雷伊布": ["サンダース", "jolteon", "雷伊布"],
    "月亮伊布": ["ブラッキー", "umbreon", "月亮伊布"],
    "仙子伊布": ["ニンフィア", "sylveon", "仙子伊布"],
}


def load_items(json_path):
    with open(json_path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    if isinstance(payload, dict):
        return payload.get("results", [])
    return payload


def fmt_jpy(value):
    return f"¥{int(value):,}" if value not in (None, "") else ""


def fmt_cny(value):
    return f"¥{value:.1f}" if value not in (None, "") else ""


def rarity_from_item(item):
    title = (item.get("title") or "").upper()
    if "PSA10" in title:
        return "PSA10"
    if "BGS10" in title:
        return "BGS10"
    if "ARS10" in title:
        return "ARS10"
    if "SAR" in title:
        return "SAR"
    if " AR" in title or "AR " in title or "AR仕様" in title:
        return "AR"
    if item.get("type") == "graded":
        return "评级卡"
    if item.get("type") == "raw":
        return "裸卡"
    return "其他"


def is_bundle(title):
    t = f" {title.lower()} "
    return any(term in t for term in NOISE_TERMS)


def section_match(section, title):
    aliases = SECTION_ALIASES.get(section, [section])
    t = title.lower()
    return any(alias.lower() in t for alias in aliases if alias)


def note_from_title(title):
    notes = []
    if "連番" in title:
        notes.append("连号")
    m = re.search(r"([2-9])枚", title)
    if m:
        notes.append(f"{m.group(1)}枚")
    if "まとめ" in title or "セット" in title:
        notes.append("组合")
    if "中国限定" in title:
        notes.append("中国限定")
    if "海外限定" in title:
        notes.append("海外限定")
    return " | ".join(notes)


def unique_rows(rows):
    seen = set()
    out = []
    for row in rows:
        key = row["url"] or (row["section"], row["title"], row["price_jpy"], row["status"])
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def normalize_items(items):
    rows = []
    for item in items:
        title = item.get("title", "").strip()
        section = item.get("section", "").strip()
        rarity = rarity_from_item(item)
        price_jpy = int(item.get("price_jpy") or 0)
        price_cny = round(price_jpy * JPY_TO_CNY, 1)
        bundle = is_bundle(title)
        relevant = section_match(section, title)
        rows.append({
            "section": section,
            "rarity": rarity,
            "type": item.get("type", ""),
            "title": title,
            "price_jpy": price_jpy,
            "price_cny": price_cny,
            "status": item.get("status", ""),
            "created": item.get("created", ""),
            "url": item.get("url", ""),
            "relevant": relevant,
            "is_bundle": bundle,
            "note": note_from_title(title),
        })

    rows = unique_rows(rows)
    add_metrics(rows)
    return rows


def add_metrics(rows):
    grouped = defaultdict(list)
    for row in rows:
        if row["relevant"] and not row["is_bundle"] and row["rarity"] in TARGET_RARITIES:
            grouped[(row["section"], row["rarity"])].append(row["price_jpy"])

    medians = {k: median(v) for k, v in grouped.items() if v}

    for row in rows:
        tags = []
        if not row["relevant"]:
            tags.append("目标不符")
        if row["is_bundle"]:
            tags.append("组合杂质")
        if row["rarity"] not in TARGET_RARITIES:
            tags.append("类型不符")
        base = medians.get((row["section"], row["rarity"]))
        row["median_jpy"] = int(base) if base else 0
        if base and row["status"] == "販売中" and row["price_jpy"] < base * 0.5:
            tags.append("疑似低价")
        row["risk_tags"] = " | ".join(tags)
        row["main_target"] = row["relevant"] and (not row["is_bundle"]) and row["rarity"] in TARGET_RARITIES
        if not row["main_target"]:
            if not row["relevant"]:
                row["noise_reason"] = "目标不符"
            elif row["is_bundle"]:
                row["noise_reason"] = "组合杂质"
            else:
                row["noise_reason"] = "类型不符"
        else:
            row["noise_reason"] = ""


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(max(width + 2, 10), 42)


def add_header(ws, title, subtitle, end_col):
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A2"] = subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)


def write_sheet(ws, rows, headers, keys, title, subtitle):
    add_header(ws, title, subtitle, len(headers))
    fill = PatternFill("solid", fgColor="D9EAF7")
    for idx, header in enumerate(headers, 1):
        c = ws.cell(row=4, column=idx, value=header)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    row_no = 5
    for row in rows:
        for col_no, key in enumerate(keys, 1):
            value = row.get(key, "")
            if key in ("price_jpy", "median_jpy"):
                value = fmt_jpy(value)
            elif key == "price_cny":
                value = fmt_cny(value)
            elif key == "url" and value:
                value = "链接"
            ws.cell(row=row_no, column=col_no, value=value)
            if key == "url" and row.get("url"):
                ws.cell(row=row_no, column=col_no).hyperlink = row["url"]
                ws.cell(row=row_no, column=col_no).style = "Hyperlink"
        row_no += 1
    ws.freeze_panes = "A5"
    autosize(ws)


def build_summary(rows):
    out = []
    for section in sorted({row["section"] for row in rows}):
        items = [x for x in rows if x["section"] == section]
        main_items = [x for x in items if x["main_target"]]
        on_sale = [x for x in main_items if x["status"] == "販売中"]
        sold_out = [x for x in main_items if x["status"] == "売り切れ"]
        lowest = min((x["price_jpy"] for x in on_sale), default=None)
        out.append({
            "section": section,
            "main_count": len(main_items),
            "on_sale": len(on_sale),
            "sold_out": len(sold_out),
            "noise_count": len(items) - len(main_items),
            "lowest_jpy": lowest,
            "lowest_cny": round(lowest * JPY_TO_CNY, 1) if lowest else None,
        })
    return out


def generate_final_report(json_path, output_path):
    rows = normalize_items(load_items(json_path))
    subtitle = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 汇率: 1 JPY = {JPY_TO_CNY} CNY | 主清单已去杂，但全量保留"

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("总览")
    write_sheet(
        ws,
        build_summary(rows),
        ["卡种", "主目标数", "在售", "售罄", "杂质数", "最低在售价JPY", "最低在售价CNY"],
        ["section", "main_count", "on_sale", "sold_out", "noise_count", "lowest_jpy", "lowest_cny"],
        "监控｜单卡｜总览",
        subtitle,
    )

    main_rows = sorted(
        [x for x in rows if x["main_target"]],
        key=lambda x: (x["section"], 0 if x["status"] == "販売中" else 1, x["rarity"], x["price_jpy"]),
    )
    ws = wb.create_sheet("主清单")
    write_sheet(
        ws,
        main_rows,
        ["卡种", "稀有度", "商品标题", "日元价格", "中位价", "人民币(≈)", "状态", "上架时间", "备注", "风险标记", "链接"],
        ["section", "rarity", "title", "price_jpy", "median_jpy", "price_cny", "status", "created", "note", "risk_tags", "url"],
        "监控｜单卡｜主清单",
        subtitle,
    )

    full_rows = sorted(rows, key=lambda x: (x["section"], x["rarity"], x["price_jpy"]))
    ws = wb.create_sheet("全量结果")
    write_sheet(
        ws,
        full_rows,
        ["卡种", "稀有度", "商品标题", "日元价格", "人民币(≈)", "状态", "上架时间", "备注", "风险标记", "链接"],
        ["section", "rarity", "title", "price_jpy", "price_cny", "status", "created", "note", "risk_tags", "url"],
        "监控｜单卡｜全量结果",
        subtitle,
    )

    noise_rows = sorted([x for x in rows if not x["main_target"]], key=lambda x: (x["section"], x["noise_reason"], x["price_jpy"]))
    ws = wb.create_sheet("杂质回收")
    write_sheet(
        ws,
        noise_rows,
        ["卡种", "杂质原因", "稀有度", "商品标题", "日元价格", "人民币(≈)", "状态", "备注", "链接"],
        ["section", "noise_reason", "rarity", "title", "price_jpy", "price_cny", "status", "note", "url"],
        "监控｜单卡｜杂质回收",
        subtitle,
    )

    wb.save(output_path)
    print(f"✅ 报告已保存: {output_path}")
    print(f"主清单: {len(main_rows)} | 全量: {len(full_rows)} | 杂质: {len(noise_rows)}")


if __name__ == "__main__":
    json_path = "/Users/hm/Desktop/蚊子報告/single_card_2026-03-17.json"
    output_path = "/Users/hm/Desktop/销售总监報告/监控_单卡_2026-03-17.xlsx"
    generate_final_report(json_path, output_path)
