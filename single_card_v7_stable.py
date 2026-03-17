#!/usr/bin/env python3
"""
每日宝可梦单卡监控报告自动化脚本
定时运行：每天 10:00 和 22:00

功能：
1. 卡号搜索（宝石包系列）
2. 售罄/在售分开显示
3. 评级卡/裸卡分流
4. 均价计算
5. 低价预警（<均价50%）
6. 生成Excel报告
7. 自动清理7天前的报告
"""

import json
import os
import sys
import time
import glob
from itertools import islice
from datetime import datetime, timedelta
from collections import defaultdict
import urllib.request
import base64
import io

# 尝试导入依赖
try:
    import mercari
    from mercari import MercariSearchStatus, MercariSort
except ImportError:
    print("❌ 缺少 mercari 库，请安装: pip install mercari-api")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ 缺少 openpyxl 库，请安装: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("❌ 缺少 requests 库，请安装: pip install requests")
    sys.exit(1)

# ============ 配置 ============
OUTPUT_DIR = os.path.expanduser("~/Desktop/销售总监報告")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
KEEP_DAYS = 7  # 保留7天报告

# 颜色定义
C_TITLE = "1A237E"
C_META = "283593"
C_WHITE = "FFFFFF"
C_RED = "FFCDD2"
C_RED_FONT = "C62828"
C_GREEN = "C8E6C9"
C_ORANGE = "FFE0B2"
C_GRADED = "FFF9C4"
C_RAW = "E3F2FD"

# Excel列定义
COL_HEADERS = ["#", "类型", "商品标题", "日元价格", "均价", "状态", "上架时间", "预警", "链接"]
COL_WIDTHS = [4, 8, 45, 12, 10, 10, 12, 8, 10]
NCOLS = len(COL_HEADERS)

# ============ 搜索关键词 ============
CARD_KEYWORDS = [
    # 原有卡组 - 宝石包/SV系列
    {"id": "sv3a_0307", "name": "イーブイ", "keywords": ["イーブイ AR", "宝石包 イーブイ", "SV3a イーブイ"]},
    {"id": "sv3a_0709", "name": "エーフィ", "keywords": ["エーフィ AR", "宝石包 エーフィ", "SV3a エーフィ"]},
    {"id": "sv3a_0115", "name": "ブラッキー", "keywords": ["ブラッキー AR", "宝石包 ブラッキー", "SV3a ブラッキー"]},
    {"id": "sv3a_0615", "name": "リーフィア", "keywords": ["リーフィア AR", "宝石包 リーフィア", "SV3a リーフィア"]},
    {"id": "sv3a_0915", "name": "グレイシア", "keywords": ["グレイシア AR", "宝石包 グレイシア", "SV3a グレイシア"]},
    {"id": "sv3a_0616", "name": "ニンフィア", "keywords": ["ニンフィア AR", "宝石包 ニンフィア", "SV3a ニンフィア"]},
    {"id": "sv4a_0410", "name": "ゲンガー", "keywords": ["ゲンガー AR", "宝石包 ゲンガー", "SV4a ゲンガー"]},
    {"id": "sv4a_0406", "name": "リザードン", "keywords": ["リザードン AR", "宝石包 リザードン", "SV4a リザードン"]},
    {"id": "sv4a_0407", "name": "カメックス", "keywords": ["カメックス AR", "宝石包 カメックス", "SV4a カメックス"]},
    {"id": "sv4a_0403", "name": "フシギダネ", "keywords": ["フシギダネ AR", "宝石包 フシギダネ", "SV4a フシギダネ"]},
    {"id": "sv5a_0809", "name": "ルギア", "keywords": ["ルギア AR", "宝石包 ルギア", "SV5a ルギア"]},
    {"id": "sv5a_0814", "name": "レックウザ", "keywords": ["レックウザ AR", "宝石包 レックウザ", "SV5a レックウザ"]},
    {"id": "sv5a_0813", "name": "サーナイト", "keywords": ["サーナイト AR", "宝石包 サーナイト", "SV5a サーナイト"]},
    {"id": "sv6a_0901_mtw", "name": "ミュウツー", "keywords": ["ミュウツー AR", "宝石包 ミュウツー", "SV6a ミュウツー"]},
    {"id": "sv6a_0901_mew", "name": "ミュウ", "keywords": ["ミュウ AR", "宝石包 ミュウ", "SV6a ミュウ"]},
    {"id": "sv6a_0902", "name": "ピカチュウ", "keywords": ["ピカチュウ AR", "宝石包 ピカチュウ", "SV6a ピカチュウ"]},
    {"id": "sv12a_1226", "name": "オーガポン", "keywords": ["オーガポン SAR", "宝石包 オーガポン", "SV12a オーガポン"]},
    
    # 新增卡号 - 中日双版并行 (卡号/XX 宽口径格式)
    {"id": "sv3a_0114", "name": "多龙梅西亚", "keywords": ["0114", "14", "ドラメिध AR", "多龙 AR", "宝石包 ドラメिध"]},
    {"id": "sv3a_0320", "name": "耿鬼ex", "keywords": ["0320", "20", "ゲンガー ex AR", "耿鬼ex AR", "宝石包 ゲンガー"]},
    {"id": "sv3a_0086", "name": "火伊布", "keywords": ["086", "66", "ブースター AR", "火伊布 AR", "宝石包 ブースター"]},
    {"id": "sv3a_0081", "name": "水伊布", "keywords": ["081", "66", "シャワーズ AR", "水伊布 AR", "宝石包 シャワーズ"]},
    {"id": "sv3a_0082", "name": "雷伊布", "keywords": ["082", "66", "サンダーの AR", "雷伊布 AR", "宝石包 サンダー"]},
    
    # 月亮伊布/仙子伊布专项 - 精准搜索 (卡号+角色名)
    {"id": "sv3a_0084", "name": "月亮伊布", "keywords": ["ブラッキー 084", "ブラッキー AR 084", "月亮伊布 AR 宝石包"]},
    {"id": "sv3a_0085", "name": "仙子伊布", "keywords": ["ニンフィア 085", "ニンフィア AR 085", "仙子伊布 AR 宝石包"]},
]

# 过滤规则
MUST_CONTAIN = ["ポケモン", "Pokémon", "Pokeca", "宝石", "ジェム", "CBB", "中国", "中国語", "海外限定", "PSA", "BGS", "ARS", "AR", "SAR"]
BLACKLIST = ["ex(", "ex　", "Gx", "VMAX", "V-UNION", "V-STAR", "VSTAR", "メガ", "無し", "セット", "まとめ売り", "引退品", "引退", "G X", "TAG TEAM", "ジャンク", "訳あり"]
NON_CARD = ["ウォレット", "バッグ", "リュック", "ワンピース", "セーター", "スマホ", "ケース", "カバー", "家電", "炊飯器", "冷蔵庫", "おもちゃ", "ゲーム", "switch", "cd", "dvd", "bluray", "リップ", "香水", "アクセサリー", "リング", "ピアス", "サングラス", "sunglasses", "zara", "ユニクロ"]

# ============ 工具函数 ============
def f(c): return PatternFill("solid", start_color=c, fgColor=c)
def bd():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)
def ca(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def la(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def col_range(row):
    return f"A{row}:{get_column_letter(NCOLS)}{row}"

def normalize(title):
    return title.replace(" ", "").replace("　", "").replace("-", "").replace(".", "").lower()

def should_include(title):
    t = title.lower()
    n = normalize(title)
    if not any(k.lower() in n for k in MUST_CONTAIN):
        return False
    for w in BLACKLIST:
        if w.lower() in t:
            return False
    for kw in NON_CARD:
        if kw.lower() in t:
            return False
    return True

def get_type(title):
    t = normalize(title)
    if "psa10" in t or "bgs10" in t or "cgc10" in t:
        return "graded"
    if ("ar" in t or "sar" in t) and not any(x in t for x in ["psa", "bgs", "cgc", "ars", "9.", "95", "9.5"]):
        return "raw"
    return None

def format_time(created):
    try:
        # 处理datetime对象
        if hasattr(created, 'year'):
            jp_time = created
        elif isinstance(created, str) and created.isdigit():
            created_dt = datetime.fromtimestamp(int(created))
            jp_time = created_dt + timedelta(hours=9)
        elif isinstance(created, str):
            created_dt = datetime.fromisoformat(created.replace('Z', '+00:00'))
            jp_time = created_dt + timedelta(hours=9)
        elif isinstance(created, (int, float)):
            created_dt = datetime.fromtimestamp(int(created))
            jp_time = created_dt + timedelta(hours=9)
        else:
            return str(created)[:10]
        
        # 返回具体日期格式
        return jp_time.strftime("%Y-%m-%d %H:%M")
    except:
        return str(created)[:10]

def get_rate():
    try:
        r = requests.get("https://wise.com/zh-cn/currency-converter/jpy-to-cny-rate", headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        import re
        m = re.search(r"¥1 JPY = ([\d.]+) CNY", r.text)
        if m: return float(m.group(1))
    except: pass
    return 0.04339

def calculate_avg_prices(items):
    price_data = defaultdict(list)
    for item in items:
        key = (item.get('section', ''), item.get('type', ''))
        if item.get('price_jpy', 0) > 0:
            price_data[key].append(item['price_jpy'])
    
    avg_prices = {}
    for key, prices in price_data.items():
        if prices:
            avg_prices[key] = {'avg': sum(prices) // len(prices), 'min': min(prices), 'count': len(prices)}
    return avg_prices

# ============ 搜索 ============
def search_card(kw_dict, max_age_days=60):
    """搜索卡片，max_age_days控制仅抓取最近N天的数据"""
    import time
    results = []
    jp = kw_dict["name"]
    
    # 计算时间 cutoff
    cutoff_ts = int(time.time()) - (max_age_days * 86400)
    
    for status in [MercariSearchStatus.ON_SALE, MercariSearchStatus.SOLD_OUT]:
        status_name = "販売中" if status == MercariSearchStatus.ON_SALE else "売り切れ"
        
        for kw in kw_dict["keywords"]:
            try:
                # 限制获取数量，避免超时
                items = list(islice(
                    mercari.search(kw, status=status, sort=MercariSort.SORT_CREATED_TIME),
                    20,
                ))
                for item in items:
                    # 检查时间过滤
                    try:
                        item_ts = int(item.created) if isinstance(item.created, (int, float)) else int(item.created) if item.created.isdigit() else 0
                        if item_ts < cutoff_ts:
                            continue  # 跳过太旧的
                    except:
                        pass
                    
                    title = item.productName
                    if not should_include(title):
                        continue
                    
                    ctype = get_type(title)
                    if not ctype:
                        continue
                    
                    results.append({
                        "title": title,
                        "price_jpy": int(item.price) if item.price else 0,
                        "status": status_name,
                        "created": format_time(item.created),
                        "created_raw": str(item.created),
                        "image_url": getattr(item, "imageURL", ""),
                        "url": item.productURL,
                        "section": jp,
                        "type": ctype,
                        "item_id": item.id
                    })
                time.sleep(0.3)
            except Exception as e:
                print(f"   ⚠️ {kw} 失败: {e}")
                time.sleep(0.5)
    
    return results

# ============ 报告生成 ============
def generate_excel(items, avg_prices, today):
    # 标记低价预警
    for item in items:
        key = (item.get('section', ''), item.get('type', ''))
        price = item.get('price_jpy', 0)
        
        if key in avg_prices and price > 0:
            avg = avg_prices[key]['avg']
            item['avg_price'] = avg
            if item.get('status') == '販売中' and price < avg * 0.5:
                item['warning'] = True
            else:
                item['warning'] = False
        else:
            item['avg_price'] = 0
            item['warning'] = False
    
    # 分类
    graded = [i for i in items if i.get('type') == 'graded']
    raw = [i for i in items if i.get('type') == 'raw']
    warnings = [i for i in items if i.get('warning')]
    
    rate = get_rate()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "监控_单卡"
    ws.sheet_view.showGridLines = False
    
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # 标题
    ws.merge_cells(col_range(1))
    c = ws["A1"]
    c.value = "监控｜单卡｜最终决策"
    c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill = f(C_TITLE); c.alignment = ca(); c.border = bd()
    ws.row_dimensions[1].height = 30
    
    # 元信息
    ws.merge_cells(col_range(2))
    c = ws["A2"]
    c.value = f"生成时间: {now_str} | 汇率: 1 JPY = {rate} CNY | 低价预警: 价格 < 均价50%"
    c.font = Font(name="Arial", size=9, color="CCCCCC")
    c.fill = f(C_META); c.alignment = la(); c.border = bd()
    ws.row_dimensions[2].height = 20
    
    ws.row_dimensions[3].height = 8
    current_row = 4
    
    # 评级卡
    ws.merge_cells(col_range(current_row))
    c = ws[f"A{current_row}"]
    c.value = "⭐ 评级卡 (PSA10/BGS10)"
    c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
    c.fill = f("F9A825"); c.alignment = la(); c.border = bd()
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    for col, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=current_row, column=col, value=h)
        c.font = Font(name="Arial", size=9, bold=True, color=C_WHITE)
        c.fill = f("546E7A"); c.alignment = ca(); c.border = bd()
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    graded_sorted = sorted(graded, key=lambda x: x.get('created_raw', '0'), reverse=True)
    for idx, item in enumerate(graded_sorted):
        r = current_row + idx
        is_warning = item.get('warning', False)
        bg = C_RED if is_warning else (C_GRADED if idx % 2 == 0 else C_WHITE)
        
        row_data = [idx + 1, "PSA10", item.get('title', ''), item.get('price_jpy', 0), item.get('avg_price', 0), item.get('status', ''), item.get('created', ''), "⚠️ 低价" if is_warning else ""]
        
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = f(bg); c.border = bd()
            if col == 3: c.alignment = la(); c.font = Font(name="Arial", size=9, color=C_RED_FONT if is_warning else "333333")
            elif col == 4: c.number_format = "¥#,##0"; c.alignment = ca(); c.font = Font(name="Arial", size=9, bold=is_warning, color=C_RED_FONT if is_warning else "333333")
            elif col == 5: c.number_format = "¥#,##0"; c.alignment = ca(); c.font = Font(name="Arial", size=9)
            elif col == 6: c.alignment = ca(); c.fill = f(C_GREEN if item.get('status') == '販売中' else C_ORANGE)
            elif col == 7: c.alignment = ca(); c.font = Font(name="Arial", size=8)
            elif col == 8: c.alignment = ca(); c.font = Font(name="Arial", size=9, bold=True, color=C_RED_FONT)
            else: c.alignment = ca(); c.font = Font(name="Arial", size=9)
        
        if item.get('url'):
            lc = ws.cell(row=r, column=9, value="🔗")
            lc.hyperlink = item['url']
            lc.font = Font(name="Arial", size=9, color="1565C0", underline="single")
            lc.alignment = ca()
            lc.fill = f(bg); lc.border = bd()
        
        ws.row_dimensions[r].height = 28
    
    current_row += len(graded_sorted) + 2
    
    # 裸卡
    ws.merge_cells(col_range(current_row))
    c = ws[f"A{current_row}"]
    c.value = "○ 未评级裸卡 (AR/SAR)"
    c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
    c.fill = f("1976D2"); c.alignment = la(); c.border = bd()
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    for col, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=current_row, column=col, value=h)
        c.font = Font(name="Arial", size=9, bold=True, color=C_WHITE)
        c.fill = f("546E7A"); c.alignment = ca(); c.border = bd()
    ws.row_dimensions[current_row].height = 20
    current_row += 1
    
    raw_sorted = sorted(raw, key=lambda x: x.get('created_raw', '0'), reverse=True)
    for idx, item in enumerate(raw_sorted):
        r = current_row + idx
        is_warning = item.get('warning', False)
        bg = C_RED if is_warning else (C_RAW if idx % 2 == 0 else C_WHITE)
        card_type = "SAR" if "sar" in item.get('title','').lower() else "AR"
        
        row_data = [idx + 1, card_type, item.get('title', ''), item.get('price_jpy', 0), item.get('avg_price', 0), item.get('status', ''), item.get('created', ''), "⚠️ 低价" if is_warning else ""]
        
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = f(bg); c.border = bd()
            if col == 3: c.alignment = la(); c.font = Font(name="Arial", size=9, color=C_RED_FONT if is_warning else "333333")
            elif col == 4: c.number_format = "¥#,##0"; c.alignment = ca(); c.font = Font(name="Arial", size=9, bold=is_warning, color=C_RED_FONT if is_warning else "333333")
            elif col == 5: c.number_format = "¥#,##0"; c.alignment = ca(); c.font = Font(name="Arial", size=9)
            elif col == 6: c.alignment = ca(); c.fill = f(C_GREEN if item.get('status') == '販売中' else C_ORANGE)
            elif col == 7: c.alignment = ca(); c.font = Font(name="Arial", size=8)
            elif col == 8: c.alignment = ca(); c.font = Font(name="Arial", size=9, bold=True, color=C_RED_FONT)
            else: c.alignment = ca(); c.font = Font(name="Arial", size=9)
        
        if item.get('url'):
            lc = ws.cell(row=r, column=9, value="🔗")
            lc.hyperlink = item['url']
            lc.font = Font(name="Arial", size=9, color="1565C0", underline="single")
            lc.alignment = ca()
            lc.fill = f(bg); lc.border = bd()
        
        ws.row_dimensions[r].height = 28
    
    current_row += len(raw_sorted) + 2
    
    # 低价预警汇总
    if warnings:
        ws.merge_cells(col_range(current_row))
        c = ws[f"A{current_row}"]
        c.value = f"⚠️ 低价预警汇总 ({len(warnings)} 件) - 价格低于均价50%"
        c.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
        c.fill = f(C_RED_FONT); c.alignment = la(); c.border = bd()
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        for idx, item in enumerate(warnings[:20]):
            r = current_row + idx
            bg = C_RED
            row_data = [idx + 1, "PSA10" if item.get('type') == 'graded' else ("SAR" if "sar" in item.get('title','').lower() else "AR"), item.get('title', '')[:40], item.get('price_jpy', 0), item.get('avg_price', 0), item.get('status', ''), "", f"{int(item.get('price_jpy',0)/item.get('avg_price',1)*100)}%"]
            
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = f(bg); c.border = bd()
                c.alignment = ca() if col != 3 else la()
                c.font = Font(name="Arial", size=9, bold=True, color=C_RED_FONT)
            
            if item.get('url'):
                lc = ws.cell(row=r, column=9, value="🔗")
                lc.hyperlink = item['url']
                lc.font = Font(name="Arial", size=9, color="1565C0", underline="single")
                lc.alignment = ca()
                lc.fill = f(bg); lc.border = bd()
    
    ws.freeze_panes = "A4"
    
    out_path = os.path.join(OUTPUT_DIR, f"监控_单卡_{today}.xlsx")
    wb.save(out_path)
    return out_path, len(graded), len(raw), len(warnings)

# ============ 清理 ============
def cleanup_old_reports():
    """删除7天前的报告"""
    pattern = os.path.join(OUTPUT_DIR, "监控_单卡_*.xlsx")
    files = glob.glob(pattern)
    cutoff = datetime.now() - timedelta(days=KEEP_DAYS)
    deleted = 0
    
    for f in files:
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(f))
            if mtime < cutoff:
                os.remove(f)
                deleted += 1
                print(f"🗑️ 删除旧报告: {os.path.basename(f)}")
        except Exception as e:
            pass
    
    return deleted

# ============ 主函数 ============
def main():
    today = datetime.now().strftime("%Y-%m-%d")
    print("=" * 60)
    print(f"监控｜单卡｜{today}")
    print("=" * 60)
    
    # 确保输出目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # 1. 搜索数据 - 默认60天
    print("\n🔍 开始搜索...")
    all_items = []
    
    for i, kw in enumerate(CARD_KEYWORDS):
        print(f"[{i+1}/{len(CARD_KEYWORDS)}] {kw['name']}...", end=" ", flush=True)
        sys.stdout.flush()
        
        # 单个卡号限时2分钟
        import signal
        
        def timeout_handler(signum, frame):
            raise TimeoutError("超时跳过")
        
        try:
            signal.signal(signal.SIGALRM, timeout_handler)
            signal.alarm(120)
            items = search_card(kw, max_age_days=7)
            signal.alarm(0)
        except (AttributeError, TimeoutError):
            items = []
            print(" ⏱️ 超时跳过", end="")
        
        print(f" {len(items)} 件")
        all_items.extend(items)
    
    # 去重
    seen = set()
    unique = [i for i in all_items if i["url"] not in seen and not seen.add(i["url"])]
    
    print(f"\n📊 总计: {len(unique)} 条数据")
    
    # 2. 保存原始JSON
    json_path = os.path.join(OUTPUT_DIR, f"single_card_{today}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"results": unique}, f, ensure_ascii=False, indent=2)
    print(f"💾 JSON保存: {json_path}")
    
    # 3. 计算均价
    avg_prices = calculate_avg_prices(unique)
    print(f"📈 计算了 {len(avg_prices)} 个类型的均价")
    
    # 4. 生成Excel
    print("\n📊 生成Excel报告...")
    out_path, graded_cnt, raw_cnt, warning_cnt = generate_excel(unique, avg_prices, today)
    print(f"✅ 已保存: {out_path}")
    
    # 5. 清理旧报告
    print("\n🧹 清理旧报告...")
    deleted = cleanup_old_reports()
    print(f"🗑️ 删除了 {deleted} 个旧报告")
    
    # 6. 统计
    on_sale = len([i for i in unique if i.get('status') == '販売中'])
    sold_out = len([i for i in unique if i.get('status') == '売り切れ'])
    
    print("\n" + "=" * 60)
    print("📊 最终统计")
    print("=" * 60)
    print(f"在售: {on_sale} | 售罄: {sold_out}")
    print(f"评级卡: {graded_cnt} | 裸卡: {raw_cnt}")
    print(f"低价预警: {warning_cnt}")
    print(f"总计: {len(unique)}")
    print("=" * 60)
    
    return out_path

if __name__ == "__main__":
    main()
