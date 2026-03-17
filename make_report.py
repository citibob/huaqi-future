#!/usr/bin/env python3
"""
煤炉监控报告生成器 v3
读取 煤炉_日期.json → 生成 煤炉报告_日期.xlsx
- Vol.1-4 分区，售罄在上/售卖在下
- 日元 + 人民币
- 上架时间列，按时间降序
- 超链接可点击
- 金色行 = 当日最低价
"""
# 先注册webp mimetype（必须在openpyxl导入之前）
import mimetypes
mimetypes.init()
if True not in mimetypes.types_map:
    mimetypes.types_map[True] = {}
mimetypes.types_map[True]['.webp'] = 'image/webp'

import json, os, re, sys, time
import base64
import urllib.request
import io
from datetime import datetime
import openpyxl
import requests
import anthropic
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from image_embed import add_image_column_mercari

# Claude API 客户端
ANTHROPIC_CLIENT = anthropic.Anthropic()

def is_pokemon_card_image(image_url: str) -> bool:
    """用Claude视觉判断图片是否是宝可梦卡牌商品"""
    if not image_url:
        return False
    try:
        req = urllib.request.Request(image_url, headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://jp.mercari.com/"
        })
        with urllib.request.urlopen(req, timeout=6) as r:
            img_bytes = r.read()
        b64 = base64.standard_b64encode(img_bytes).decode()
        # 判断Content-Type
        ext = image_url.split(".")[-1].split("?")[0].split("@")[0]
        media_type = "image/jpeg" if ext in ["jpg","jpeg"] else "image/webp" if ext == "webp" else "image/png"
        resp = ANTHROPIC_CLIENT.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=10,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                    {"type": "text", "text": "Is this a Pokemon card or Pokemon card product? Reply only YES or NO."}
                ]
            }]
        )
        return "YES" in resp.content[0].text.upper()
    except Exception:
        return True  # 下载失败默认保留

OUTPUT_DIR = os.path.expanduser("~/Desktop/销售总监報告")

# ── 颜色 ──
C_TITLE      = "0D1B2A"
C_META       = "1B2838"
C_VOL_HDR    = "1A237E"
C_SOLD_HDR   = "7B1FA2"
C_SOLD_COL   = "9C27B0"
C_SOLD_EVEN  = "F3E5F5"
C_SOLD_ODD   = "FFFFFF"
C_SALE_HDR   = "1B5E20"
C_SALE_COL   = "2E7D32"
C_SALE_EVEN  = "E8F5E9"
C_SALE_ODD   = "FFFFFF"
C_WHITE      = "FFFFFF"
C_GOLD       = "FFD600"
C_GOLD_DARK  = "F57F17"

COL_HEADERS = ["#", "状态", "关键词", "日语标题", "日元价格", "人民币(≈)", "备注", "商品链接", "上架时间", "图片"]
COL_WIDTHS  = [4,   7,     20,      48,         13,         13,          16,       16,         14,        8]
NCOLS = len(COL_HEADERS)

def f(c):
    return PatternFill("solid", start_color=c, fgColor=c)

def bd(color="DDDDDD"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def ca(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def la(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def time_ago(ts_str):
    try:
        # 处理datetime对象
        if hasattr(ts_str, 'year'):
            return ts_str.strftime("%Y-%m-%d %H:%M")
        
        ts = int(ts_str)
        # 返回具体日期时间格式
        dt = datetime.fromtimestamp(ts)
        # 转换为日本时间 (UTC+9)
        dt_jp = dt.replace(hour=dt.hour + 9)
        if dt_jp.day != dt.day:
            dt_jp = dt_jp.replace(day=dt.day + 1)  # 简单处理跨日
        return dt_jp.strftime("%Y-%m-%d %H:%M")
    except:
        return "—"

# ── 宝可梦卡片名对照 ──
CARD_NAMES = {
    "カラカル":"狞猫","ゲンガー":"耿鬼","ソウブレイズ":"霸主剑","ニャース":"喵喵",
    "リザードン":"喷火龙","ピカチュウ":"皮卡丘","イーブイ":"伊布","ミュウ":"梦幻",
    "ミュウツー":"超梦","シャワーズ":"水伊布","サンダース":"雷伊布","ブースター":"火伊布",
    "エーフィ":"太阳伊布","ブラッキー":"月亮伊布","グレイシア":"冰伊布",
    "リーフィア":"叶伊布","ニンフィア":"仙子伊布","ルカリオ":"路卡利欧",
    "ガブリアス":"烈咬陆鲨","カイリュー":"快龙","カビゴン":"卡比兽",
}

def parse_note(title: str) -> str:
    notes = []
    # BOX数量
    bm = re.search(r'(\d+)\s*(?:box|ボックス|箱|BOX)', title, re.IGNORECASE)
    if bm: notes.append(f"{bm.group(1)}盒")
    # カートン
    if "カートン" in title:
        cm = re.search(r'(\d+)\s*カートン', title)
        notes.append(f"{cm.group(1)}カートン" if cm else "1カートン")
    # 包数
    pm = re.search(r'(\d+)\s*(?:パック|pack)', title, re.IGNORECASE)
    if pm: notes.append(f"{pm.group(1)}包")
    # Vol混合
    vols = sorted(set(re.findall(r'[Vv]ol[.\s]*(\d)', title)))
    if len(vols) >= 2: notes.append(f"Vol.{'.'.join(vols)}混合")
    # 单卡 PSA/BGS
    if re.search(r'PSA\s*\d+|BGS\s*[\d.]+', title, re.IGNORECASE):
        gm = re.search(r'PSA\s*(\d+)|BGS\s*([\d.]+)', title, re.IGNORECASE)
        g = gm.group(1) or gm.group(2)
        card_jp = next((k for k in CARD_NAMES if k in title), None)
        card_cn = CARD_NAMES.get(card_jp, card_jp) if card_jp else ""
        prefix = "PSA" if "PSA" in title.upper() else "BGS"
        notes.append(f"单卡·{card_cn}({prefix}{g})" if card_cn else f"单卡({prefix}{g})")
    # 单卡稀有度
    elif re.search(r'\b(SAR|AR|SR|UR|RR)\b', title):
        card_jp = next((k for k in CARD_NAMES if k in title), None)
        card_cn = CARD_NAMES.get(card_jp, card_jp) if card_jp else "稀有卡"
        rar = re.search(r'\b(SAR|AR|SR|UR|RR)\b', title).group(1)
        notes.append(f"单卡·{card_cn}({rar})")
    # 兜底
    if not notes:
        notes.append("1盒(未開封)" if ("未開封" in title or "新品" in title) else "—")
    return " / ".join(notes)

def get_rate():
    try:
        r = requests.get("https://wise.com/zh-cn/currency-converter/jpy-to-cny-rate",
                         headers={"User-Agent":"Mozilla/5.0"}, timeout=10)
        m = re.search(r"¥1 JPY = ([\d.]+) CNY", r.text)
        if m: return float(m.group(1))
    except: pass
    return 0.04339

def col_range(row):
    return f"A{row}:{get_column_letter(NCOLS)}{row}"

def write_section(ws, start_row, label, items, is_on_sale, rate, items_with_rows):
    hdr_c  = C_SALE_HDR  if is_on_sale else C_SOLD_HDR
    col_c  = C_SALE_COL  if is_on_sale else C_SOLD_COL
    even_c = C_SALE_EVEN if is_on_sale else C_SOLD_EVEN
    icon   = "🟢" if is_on_sale else "🔴"

    # 区块标题
    ws.merge_cells(col_range(start_row))
    c = ws[f"A{start_row}"]
    c.value = f"   {icon}  {label}  （{len(items)} 件）"
    c.font = Font(name="Arial", size=12, bold=True, color=C_WHITE)
    c.fill = f(hdr_c)
    c.alignment = la()
    c.border = bd("888888")
    ws.row_dimensions[start_row].height = 26
    header_rows = {start_row}  # 记录非数据行
    start_row += 1

    # 列标题
    for col, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = Font(name="Arial", size=9, bold=True, color=C_WHITE)
        c.fill = f(col_c)
        c.alignment = ca()
        c.border = bd()
    ws.row_dimensions[start_row].height = 22
    header_rows.add(start_row)
    start_row += 1

    if not items:
        ws.merge_cells(col_range(start_row))
        c = ws[f"A{start_row}"]
        c.value = "  （データなし）"
        c.font = Font(name="Arial", size=9, italic=True, color="999999")
        c.fill = f(even_c)
        c.alignment = la()
        c.border = bd()
        ws.row_dimensions[start_row].height = 20
        return start_row + 1

    # 过滤30天以前的数据
    cutoff = int(time.time()) - 86400 * 60
    items = [it for it in items if int(it.get("created","0") or "0") >= cutoff]

    # 用Claude视觉API验证图片是否是宝可梦卡牌
    print(f"🔍 开始图片验证 ({len(items)} 条数据)...")
    validated_items = []
    for it in items:
        if is_pokemon_card_image(it.get("image_url", "")):
            validated_items.append(it)
        else:
            print(f"🔍 跳过非卡牌: {it.get('title', '')[:40]}")
    print(f"🔍 验证完成: {len(validated_items)}/{len(items)} 条是卡牌")
    items = validated_items

    # 按时间降序排列
    sorted_items = sorted(items,
        key=lambda x: int(x.get("created","0") or "0"), reverse=True)

    # 找最低价索引（仅售卖区）
    cheapest_idx = None
    if is_on_sale:
        prices = [it.get("price", 999999) for it in sorted_items]
        cheapest_idx = prices.index(min(prices)) if prices else None

    for i, item in enumerate(sorted_items):
        r = start_row + i
        is_cheap = (i == cheapest_idx)
        bg = C_GOLD if is_cheap else (even_c if i % 2 == 0 else C_SOLD_ODD)
        jpy = item.get("price", 0)
        cny = round(int(jpy) * rate, 1)
        created = item.get("created", "")

        row_data = [
            i + 1,
            "販売中" if is_on_sale else "売り切れ",
            item.get("keyword", ""),
            item.get("title", ""),
            jpy,
            cny,
            parse_note(item.get("title", "")),
            None,
            time_ago(created),
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = f(bg)
            c.border = bd()
            bold = is_cheap

            if col == 5:
                c.value = jpy
                c.number_format = "¥#,##0"
                c.font = Font(name="Arial", size=9, bold=bold, color="000000")
                c.alignment = ca()
            elif col == 6:
                c.value = cny
                c.number_format = '#,##0.0"元"'
                c.font = Font(name="Arial", size=9, bold=bold, color="000000")
                c.alignment = ca()
            elif col in [1, 2, 9]:
                c.font = Font(name="Arial", size=9, bold=bold, color="000000")
                c.alignment = ca()
            elif col == 4:
                c.font = Font(name="Arial", size=9, bold=bold, color="000000")
                c.alignment = la()
                ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 30)
            else:
                c.font = Font(name="Arial", size=9, bold=bold, color="000000")
                c.alignment = la()

        # 超链接
        url = item.get("url", "")
        lc = ws.cell(row=r, column=8, value="🔗 查看" if url else "—")
        if url:
            lc.hyperlink = url
            lc.font = Font(name="Arial", size=9, color="1565C0", underline="single", bold=is_cheap)
        else:
            lc.font = Font(name="Arial", size=9, color="AAAAAA")
        lc.fill = f(bg)
        lc.border = bd()
        lc.alignment = ca()

        if not ws.row_dimensions[r].height:
            ws.row_dimensions[r].height = 22

        # 记录图片嵌入需要的行号和item
        items_with_rows.append((r, item))

    return start_row + len(sorted_items), header_rows


def main():
    today = datetime.now().strftime("%Y-%m-%d")
    json_path = os.path.join(OUTPUT_DIR, f"煤炉_{today}.json")

    if not os.path.exists(json_path):
        print(f"❌ 找不到：{json_path}")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as fh:
        data = json.load(fh)

    results  = data.get("results", [])
    on_sale  = [r for r in results if r.get("status") == "販売中"]
    sold_out = [r for r in results if r.get("status") == "売り切れ"]
    print(f"📊 販売中 {len(on_sale)} | 売り切れ {len(sold_out)}")

    rate = get_rate()
    print(f"💱 1 JPY = {rate} CNY")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "煤炉监控"
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 主标题
    ws.merge_cells(col_range(1))
    c = ws["A1"]
    c.value = "监控｜煤炉宝石包"
    c.font = Font(name="Arial", size=14, bold=True, color=C_WHITE)
    c.fill = f(C_TITLE)
    c.alignment = ca()
    c.border = bd("000000")
    ws.row_dimensions[1].height = 38

    # 元信息
    ws.merge_cells(col_range(2))
    c = ws["A2"]
    c.value = (f"  生成时间：{now_str}　|　"
               f"汇率：1 JPY = {rate} CNY（Wise {today}）　|　"
               f"数据来源：jp.mercari.com")
    c.font = Font(name="Arial", size=8, color="CCCCCC")
    c.fill = f(C_META)
    c.alignment = la()
    c.border = bd("000000")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    current_row = 4
    items_with_rows = []  # 用于图片嵌入
    all_header_rows = set()  # 记录所有非数据行

    for vol in ["vol.1", "vol.2", "vol.3", "vol.4"]:
        vol_label = vol.upper()

        # Vol 大标题
        ws.merge_cells(col_range(current_row))
        c = ws[f"A{current_row}"]
        c.value = f"   📦   {vol_label}"
        c.font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
        c.fill = f(C_VOL_HDR)
        c.alignment = la()
        c.border = bd("000000")
        ws.row_dimensions[current_row].height = 32
        all_header_rows.add(current_row)  # Vol标题行
        current_row += 1

        vol_sold = [r for r in sold_out if vol in r.get("keyword","").lower()]
        vol_sale = [r for r in on_sale  if vol in r.get("keyword","").lower()]

        current_row, header_rows = write_section(ws, current_row, "売り切れ（已售罄 · 参考价格）", vol_sold, False, rate, items_with_rows)
        all_header_rows.update(header_rows)
        ws.row_dimensions[current_row].height = 4
        current_row += 1
        current_row, header_rows = write_section(ws, current_row, "販売中（贩卖中）", vol_sale, True, rate, items_with_rows)
        all_header_rows.update(header_rows)

        ws.row_dimensions[current_row].height = 14
        current_row += 1

    # 底部
    current_row += 1
    ws.merge_cells(col_range(current_row))
    c = ws[f"A{current_row}"]
    c.value = f"  ※ 人民币为估算值（Wise {today}）。🟡 金色行 = 该Vol当日最低价。上架时间按最新排序。"
    c.font = Font(name="Arial", size=8, italic=True, color="888888")
    c.alignment = la()
    ws.row_dimensions[current_row].height = 16

    ws.freeze_panes = "A4"

    # 嵌入图片
    print("🖼️ 正在嵌入图片...")
    add_image_column_mercari(ws, items_with_rows, all_header_rows)

    out_path = os.path.join(OUTPUT_DIR, f"监控_煤炉宝石包_{today}.xlsx")
    
    # ========== 彻底修复 webp mimetype 问题 ==========
    import mimetypes
    import sys
    
    # 方案1: 修复全局 mimetypes
    mimetypes.init()
    mimetypes.types_map.setdefault('.webp', 'image/webp')
    if True not in mimetypes.types_map:
        mimetypes.types_map[True] = {}
    mimetypes.types_map[True].setdefault('.webp', 'image/webp')
    
    # 方案2: patch openpyxl Manifest._register_mimetypes
    from openpyxl.packaging import manifest as manifest_module
    _orig_register = manifest_module.Manifest._register_mimetypes
    
    def _patched_register(self, filenames):
        for fn in filenames:
            ext = os.path.splitext(fn)[-1]
            if not ext:
                continue
            # 使用 get 避免 KeyError
            mime = mimetypes.types_map.get(True, {}).get(ext) or mimetypes.types_map.get(ext)
            if mime is None:
                # 如果找不到mime type，使用默认值
                mime = 'application/octet-stream'
            from openpyxl.packaging.manifest import FileExtension
            fe = FileExtension(ext[1:], mime)
            self.Default.append(fe)
    
    manifest_module.Manifest._register_mimetypes = _patched_register
    wb.save(out_path)
    print(f"✅ 已保存：{out_path}")

if __name__ == "__main__":
    main()
