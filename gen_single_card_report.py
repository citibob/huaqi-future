#!/usr/bin/env python3
"""
单卡监控报告生成器 v2 - 严格过滤高价值宝可梦卡片
"""
import json
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================================================================
# 配置
# ================================================================
timestamp = "2026-03-19"
json_path = f"/Users/hm/Desktop/销售总监報告/single_card_{timestamp}.json"
output_path = f"/Users/hm/Desktop/销售总监報告/监控_单卡_{timestamp}.xlsx"

with open(json_path) as f:
    data = json.load(f)
results = data["results"]

# ================================================================
# 辅助函数
# ================================================================
def norm(title):
    """normalize title for matching"""
    return title.replace(" ", "").replace("　", "").replace("-", "").replace(".", "").lower()

# ================================================================
# 第一层过滤：确认是 Pokemon TCG 商品
# ================================================================

# 必须有以下任一，才算 Pokemon TCG
POKEMON_CORE_KW = [
    # 直接 Pokemon 卡标识
    "ポケモン", "Pokemon", "Pokémon", "Pokeca", " pokeca", "トレカ",
    # 评级
    "psa", "bgs", "cgc",
    # 卡种
    "sar", " ar ", "/ar", "/ar/", "ar/", " ar", "ar ",
    "ssr", "ex/", "/ex", "ex ",
    # 卡包/系列
    "sv3a", "sv4a", "sv5a", "sv6a", "sv12a", "sv1a", "sv2a",
    "sv7a", "sv8a", "sv9a",
    "gem pack", "gem-pack", "ジェムパック", "宝石包",
    "Alternate Art", "alternateart",
    "shining", "shiny",
    # 特定宝可梦名称（日英双语）
    " pikachu", "pikachu", "皮卡丘",
    " eevee", "eevee", "evee", "イーブイ",
    " umbreon", "umbr",
    " espeon", "espec",
    " leafeon", "leaf",
    " glaceon", "glac",
    " sylveon", "sylv",
    " gardevoir", "サーナイト",
    " lugia", "ルギア",
    " rayquaza", "レックウザ",
    " mew", "ミュウ",
    " mewtwo", "ミュウツー",
    " gengar", "ゲンガー",
    " charizard", "リザードン",
    " blastoise", "カメックス",
    " venusaur", "フシギダネ",
    " garchomp",
    " arceus", "アルセウス",
    " ogerpon", "オーガポン",
    " ゲッコウガ",
    " スイレン",
    " ガチグマ",
    " サザレ",
    " ナンジャモ",
    " ミミッキュ",
    " リオル",
    " ルカリオ",
    " メタモン",
    " CAC", " CBB",
    " 中国語", "中国版", "海外限定",
    " クリムゼ", " Crims",
    " AR013", "AR013",
    " AR012", "AR012",
    " 151/", "151/",
    " cla ", " CLA ", "clair",
]

# 非 Pokemon 商品的明确标识
NON_POKEMON = [
    # GPS/雷达
    "AR-4", "AR-47", "AR47", "AR-3", "AR-1",
    # 服装/时尚
    "ARROWS", " jacket", " Jacket", " jacket", " Coat", " coat",
    " blazer", " Blazer", " pants", " Pants", " jeans", " Jeans",
    " dress", " Dress", " skirt", " Skirt", " shirt", " Shirt",
    " sweater", " Sweater", " knit", " Knit", "cardigan", "Cardigan",
    " skirt", " Skirt", " culottes", " pants",
    " stuart", "Stuart", " Paul", "MARK&LONA", "LONA",
    " ARNOTT", "Arnott",
    " comdes", "comme des garcons", " cdg",
    " zara", "uniqlo", " GU", "、無印",
    # 手表/首饰
    " QUARTZ", "quartz", " SEIKO", "seiko", " watch", " Watch",
    " OMEGA", "omega", " ROLEX", "rolex",
    " BVLGARI", "bvlgari", " CARTIER", "cartier",
    # 汽车/模型
    " MERCEDES", "mercedes", " F1", " mini", " MINICHAMPS",
    " 1/43", "1/43", " golf", " Golf",
    # 电子产品
    "iPhone", " iphone", " iphone", " Android", " android",
    " carplay", "CarPlay", " androidauto", "AndroidAuto",
    " ドライブレコーダー", " 行車記錄儀",
    "エレコム", " SANWA", "sanwa",
    # 玩具/非TCG
    "プラモデル", " プラ", " fig", " Fig",
    " ウマ娘", " 咒術", " エヴァ", "エヴァ", " 五条",
    " ブルマァク", " M1号",
    # 娱乐
    " NCT ", " BTS ", " ARASHI", "Arashi", " ODYSSEY",
    # 家居
    " Gardenmaster", "gardenmaster", " NITTO", " Nitto",
    # 其他
    " child", " Child", " kids", " Kids",
    " 男の子", " 女の子", " baby", " Baby",
    " size", " Size",
]

def nkw(kw):
    """关键词 normalize（与 title 相同的 normalize）"""
    return kw.replace(" ", "").replace("　", "").replace("-", "").replace(".", "").lower()

def is_pokemon_tcg(item_title):
    """第一层：确认是 Pokemon TCG 相关商品"""
    t = item_title.lower()
    n = norm(item_title)

    # 核心 Pokemon 关键词检测（原始title和normalized title都要匹配）
    has_core = any(kw.lower() in t or nkw(kw) in n for kw in POKEMON_CORE_KW)
    if not has_core:
        return False

    # 明确非 Pokemon 检测
    for junk in NON_POKEMON:
        if junk.lower() in t or junk.lower() in n:
            return False

    return True

# ================================================================
# 稀有度分类
# ================================================================
def classify_rarity(title):
    """分类稀有度"""
    n = norm(title)
    t = title.lower()

    # PSA/BGS/CGC 优先
    if "psa10" in n or "bgs10" in n or "cgc10" in n:
        return "PSA10"
    if "psa9" in n or "bgs9" in n or "cgc9" in n:
        return "PSA9"
    if "psa8" in n or "bgs8" in n or "cgc8" in n:
        return "PSA8"
    if "psa7" in n or "bgs7" in n or "cgc7" in n:
        return "PSA7"

    # SAR（需要 Pokemon 上下文）
    if "sar" in n:
        # 检查 Pokemon 上下文
        pokemon_context = any(kw in t or kw in n for kw in [
            " pokemon", " sv", " PSA", "psa", " ex", "ssr",
            " ゲコ", " スイレン", " ガチグマ", " サーナイト",
            " ナージャ", " オーガポン", " ゲッコウガ",
            " pikachu", " charizard", " lugia", " mew",
            " gardevoir", " arceus", " umbreon",
            " Miraidon", " Walking", " Choci",
            " クレイバー", " アクジキ",
            " クリムゼ", " Cress",
            " CRI", " CRI-",
            " CAC", " CBB",
            " 中国", " 海外",
            " アルカナディア", " アルアリル",
        ])
        if pokemon_context:
            return "SAR"
        # 独立 SAR（非 Pokemon）
        return "OTHER"

    # AR（需要 Pokemon 上下文）
    # AR 在 normalized 后需要单独检查（因为 Umbreon 里有 ar）
    if " ar " in f" {t} " or " ar/" in n or "/ar" in n or "/ar/" in n or "ar " in n[-4:]:
        # Pokemon 上下文
        pokemon_context = any(
            (kw in t or nkw(kw) in n)
            for kw in [
                " pokemon", " PSA", "psa", " SSR",
                " sv", " SV",
                " アルセウス",
                " CAC", " CBB",
                " 中国", " 海外", "umbreon", "espeon",
                " AR013", " AR012",
                " クリムゼ", " Crims", " CRI",
            ]
        )
        if pokemon_context:
            return "AR"
        return "OTHER"

    # SSR / ex
    if "ssr" in n:
        return "SSR"
    if "ex" in n and (" pokemon" in t or " sv" in n or " ポケモン" in t):
        return "ex"

    # 无稀有度标识
    return "OTHER"

# ================================================================
# 价值筛选
# ================================================================
def is_valuable(item):
    """高价值筛选"""
    title = item["title"]
    rarity = classify_rarity(title)
    price = item["price_jpy"]

    if rarity in ["PSA10", "PSA9", "PSA8", "PSA7"]:
        if rarity == "PSA10" and price >= 3000:
            return True
        if rarity == "PSA9" and price >= 2000:
            return True
        if rarity in ["PSA8", "PSA7"] and price >= 1000:
            return True
        return False

    if rarity == "SAR" and price >= 2000:
        return True
    if rarity == "AR" and price >= 2000:
        return True
    if rarity == "SSR" and price >= 3000:
        return True
    if rarity == "ex" and price >= 3000:
        return True

    return False

# ================================================================
# 执行过滤
# ================================================================
filtered = []
for item in results:
    title = item["title"]
    if is_pokemon_tcg(title):
        item["rarity"] = classify_rarity(title)
        filtered.append(item)

print(f"原始: {len(results)} 件 → Pokemon TCG 相关: {len(filtered)} 件")

valuable = [i for i in filtered if is_valuable(i)]
print(f"高价值: {len(valuable)} 件")

# Debug
print("\n[Pokemon TCG 相关卡片]:")
for i, item in enumerate(filtered):
    r = classify_rarity(item["title"])
    val = is_valuable(item)
    print(f"  [{i+1:2d}] {item.get('section','?'):8s} | {r:6s} | {'✅' if val else '❌'} | ¥{item['price_jpy']:>8} | {item['title'][:50]}")

# 按 section 分组
sections = {}
for item in valuable:
    sec = item.get("section", "其他")
    if sec not in sections:
        sections[sec] = []
    sections[sec].append(item)

print("\n[按卡种分组]:")
for sec, items in sorted(sections.items()):
    on_sale = [i for i in items if i["status"] == "販売中"]
    sold = [i for i in items if i["status"] == "売り切れ"]
    print(f"  {sec}: 在售{len(on_sale)}/售罄{len(sold)}")

# ================================================================
# 生成 Excel
# ================================================================
wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GRADED_FILL = PatternFill("solid", fgColor="FFF2CC")
RAW_FILL = PatternFill("solid", fgColor="E2EFDA")
SOLD_FILL = PatternFill("solid", fgColor="FCE4D6")
ON_SALE_FILL = PatternFill("solid", fgColor="DDEBF7")
BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

def style_cell(cell, fill=None, bold=False, align='left', wrap=False, size=10):
    if fill:
        cell.fill = fill
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = BORDER

# Sheet 1: 总览
ws1 = wb.active
ws1.title = "总览"
headers = ["序号", "卡种", "稀有度", "状态", "价格(円)", "上架时间", "标题", "链接"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=col, value=h)
    style_cell(c, fill=HEADER_FILL, bold=True, align='center')
ws1.row_dimensions[1].height = 22

row = 2
for i, item in enumerate(valuable, 1):
    rarity = item.get("rarity", item.get("type", "OTHER"))
    title = item["title"]
    status = item["status"]

    if "psa10" in title.lower() or "psa 10" in title.lower():
        row_fill = GRADED_FILL
    elif "psa" in title.lower() or "bgs" in title.lower():
        row_fill = PatternFill("solid", fgColor="FFFDE7")
    else:
        row_fill = RAW_FILL

    vals = [i, item.get("section", ""), rarity, status, item["price_jpy"], item["created"], title, item["url"]]
    for col, v in enumerate(vals, 1):
        c = ws1.cell(row=row, column=col, value=v)
        style_cell(c, fill=row_fill, align='center' if col in [1,3,4,5] else 'left', wrap=(col==7))
    ws1.row_dimensions[row].height = 30
    row += 1

ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 8
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 16
ws1.column_dimensions['G'].width = 50
ws1.column_dimensions['H'].width = 35
ws1.freeze_panes = 'A2'

# Sheet 2: 按卡种详细
ws2 = wb.create_sheet("按卡种")

def write_section(ws, section_name, items, start_row):
    r = start_row
    c_title = ws.cell(row=r, column=1, value=f"【{section_name}】")
    c_title.font = Font(bold=True, size=12, color="FFFFFF")
    c_title.fill = PatternFill("solid", fgColor="2E75B6")
    c_title.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 22
    r += 1

    hdrs = ["序号", "稀有度", "状态", "价格(円)", "上架时间", "标题", "链接"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=col, value=h)
        style_cell(c, fill=PatternFill("solid", fgColor="BDD7EE"), bold=True, align='center')
    ws.row_dimensions[r].height = 18
    r += 1

    on_sale = [i for i in items if i["status"] == "販売中"]
    sold = [i for i in items if i["status"] == "売り切れ"]

    for i, item in enumerate(on_sale, 1):
        title = item["title"]
        if "psa10" in title.lower():
            f = GRADED_FILL
        else:
            f = PatternFill("solid", fgColor="DEEAF1")
        rarity = item.get("rarity", item.get("type", "OTHER"))
        vals = [i, rarity, "販売中", item["price_jpy"], item["created"], title, item["url"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            style_cell(c, fill=f, align='center' if col in [1,2,3,4] else 'left', wrap=(col==6))
        ws.row_dimensions[r].height = 28
        r += 1

    for i, item in enumerate(sold, len(on_sale)+1):
        title = item["title"]
        if "psa10" in title.lower():
            f = GRADED_FILL
        else:
            f = SOLD_FILL
        rarity = item.get("rarity", item.get("type", "OTHER"))
        vals = [i, rarity, "売り切れ", item["price_jpy"], item["created"], title, item["url"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            style_cell(c, fill=f, align='center' if col in [1,2,3,4] else 'left', wrap=(col==6))
        ws.row_dimensions[r].height = 28
        r += 1

    c_sum = ws.cell(row=r, column=1, value=f"  小计：在售{len(on_sale)}件 / 售罄{len(sold)}件")
    c_sum.font = Font(bold=True, size=10, italic=True)
    c_sum.fill = PatternFill("solid", fgColor="E7E6E6")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.row_dimensions[r].height = 16
    r += 2
    return r

r = 1
for sec in sorted(sections.keys()):
    items = sections[sec]
    r = write_section(ws2, sec, items, r)

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 9
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 11
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 55
ws2.column_dimensions['G'].width = 35

# Sheet 3: 统计摘要
ws3 = wb.create_sheet("统计摘要")
summary_data = [
    ["单卡监控统计摘要", ""],
    ["监控日期", timestamp],
    ["原始抓取总数", len(results)],
    ["Pokemon TCG相关", len(filtered)],
    ["高价值筛选后", len(valuable)],
    ["", ""],
    ["—— 按稀有度分布 ——", ""],
    ["PSA10", len([i for i in valuable if i.get("rarity") == "PSA10"])],
    ["PSA9", len([i for i in valuable if i.get("rarity") == "PSA9"])],
    ["PSA8", len([i for i in valuable if i.get("rarity") == "PSA8"])],
    ["SAR (裸)", len([i for i in valuable if i.get("rarity") == "SAR"])],
    ["AR (裸)", len([i for i in valuable if i.get("rarity") == "AR"])],
    ["SSR", len([i for i in valuable if i.get("rarity") == "SSR"])],
    ["ex", len([i for i in valuable if i.get("rarity") == "ex"])],
    ["", ""],
    ["—— 按状态分布 ——", ""],
    ["在售中", len([i for i in valuable if i["status"] == "販売中"])],
    ["售罄", len([i for i in valuable if i["status"] == "売り切れ"])],
]

for i, (col1, col2) in enumerate(summary_data, 1):
    c1 = ws3.cell(row=i, column=1, value=col1)
    c2 = ws3.cell(row=i, column=2, value=col2)
    if i == 1:
        c1.font = Font(bold=True, size=14)
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    elif col1.startswith("——"):
        c1.font = Font(bold=True, color="FFFFFF")
        c1.fill = PatternFill("solid", fgColor="2E75B6")
        ws3.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    else:
        c1.font = Font(bold=True)
        c2.font = Font(size=12)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15

wb.save(output_path)
print(f"\n✅ Excel报告已生成: {output_path}")
