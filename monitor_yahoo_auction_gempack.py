#!/usr/bin/env python3
import html
import json
import os
import re
from collections import OrderedDict, defaultdict
from datetime import datetime
from statistics import median
from urllib.parse import quote_plus

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT_DIR = os.path.expanduser("~/Desktop/销售总监報告")
OUT_FILE = os.path.join(OUTPUT_DIR, f"监控_雅虎拍卖宝石包_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
CACHE_FILE = os.path.expanduser(f"~/.openclaw/agents/ebay/yahoo_gem_monitor_{datetime.now().strftime('%Y-%m-%d')}.json")
JPY_TO_CNY = 0.04334
VOLS = ["VOL.1", "VOL.2", "VOL.3", "VOL.4"]
SEARCH_MATRIX = OrderedDict({
    "VOL.1": ["宝石包 vol.1", "宝石包 第一弹", "ジェムパック vol.1", "gem pack cbb1c"],
    "VOL.2": ["宝石包 vol.2", "宝石包 第二弹", "ジェムパック vol.2", "gem pack cbb2c"],
    "VOL.3": ["宝石包 vol.3", "宝石包 第三弹", "ジェムパック vol.3", "gem pack cbb3c"],
    "VOL.4": ["宝石包 vol.4", "宝石包 第四弹", "ジェムパック vol.4", "gem pack cbb4c"],
})
BOX_TERMS = ("box", "ボックス", "箱", "未開封", "シュリンク")
PACK_TERMS = ("pack", "パック", "10パック", "1パック")
SINGLE_TERMS = ("psa", "bgs", "ミラー", "シングル")
CORE_TERMS = ("宝石包", "ジェムパック", "gem pack", "gempack")


def fmt_jpy(v):
    return f"¥{int(v):,}" if isinstance(v, (int, float)) and v else ""


def fmt_cny(v):
    return f"¥{v:.1f}" if isinstance(v, (int, float)) and v is not None else ""


def detect_type(title: str):
    t = title.lower()
    if re.search(r"\b(?:sar|sr|ar|rr|ur)\b", t):
        return "单卡"
    if any(x in t for x in SINGLE_TERMS):
        return "单卡"
    if any(x in t for x in PACK_TERMS):
        return "散包"
    if any(x in t for x in BOX_TERMS):
        return "整盒"
    return "未分类"


def detect_vol(title: str):
    t = title.upper()
    m = re.search(r"VOL[.\s]*([1-4])", t)
    if m:
        return f"VOL.{m.group(1)}"
    m = re.search(r"CBB([1-4])C", t)
    if m:
        return f"VOL.{m.group(1)}"
    if "第一弹" in title or "第1弹" in title:
        return "VOL.1"
    if "第二弹" in title or "第2弹" in title:
        return "VOL.2"
    if "第三弹" in title or "第3弹" in title:
        return "VOL.3"
    if "第四弹" in title or "第4弹" in title:
        return "VOL.4"
    return ""


def relevant_title(title: str, query_vol: str):
    t = title.lower()
    has_core = any(x in t for x in CORE_TERMS)
    has_cbb = bool(re.search(r"\bcbb[1-4]c\b", t))
    has_pokemon = ("ポケモン" in title) or ("pokemon" in t)
    if not has_core and not (has_cbb and has_pokemon):
        return False
    detected = detect_vol(title)
    return bool(detected and detected == query_vol)


def note_from_title(title: str):
    t = title.lower()
    notes = []
    m = re.search(r"(\d+)\s*(?:box|箱)", t)
    if m:
        notes.append(f"{m.group(1)}盒")
    m = re.search(r"(\d+)\s*(?:pack|パック)", t)
    if m:
        notes.append(f"{m.group(1)}包")
    return " | ".join(notes)


def safe_int(text: str):
    if not text:
        return None
    m = re.search(r"([\d,]+)", text)
    return int(m.group(1).replace(",", "")) if m else None


def search_urls():
    session = requests.Session()
    session.headers.update({"user-agent": "Mozilla/5.0"})
    out = OrderedDict()
    for vol, keywords in SEARCH_MATRIX.items():
        for keyword in keywords:
            url = f"https://auctions.yahoo.co.jp/search/search?p={quote_plus(keyword)}"
            try:
                text = session.get(url, timeout=20).text
            except Exception:
                continue
            for href in re.findall(r'href="(https://auctions\.yahoo\.co\.jp/jp/auction/[^"]+)"', text):
                href = html.unescape(href)
                if href not in out:
                    out[href] = {"query_vol": vol, "keyword": keyword}
    return out


def load_cache_rows():
    if not os.path.exists(CACHE_FILE):
        return []
    with open(CACHE_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("rows", [])


def parse_detail(session, url: str, query_vol: str, keyword: str):
    text = session.get(url, timeout=20).text
    title = html.unescape(re.search(r"<title>(.*?)</title>", text, re.S).group(1)).split(" - Yahoo!オークション")[0].strip()
    if not relevant_title(title, query_vol):
        return None
    price_blocks = re.findall(r"(?:現在|即決)[^0-9]{0,20}([\d,]+)円", text)
    current_price = safe_int(price_blocks[0]) if price_blocks else None
    buyout_match = re.search(r"即決[^0-9]{0,20}([\d,]+)円", text)
    buyout_price = safe_int(buyout_match.group(1)) if buyout_match else None
    bid_match = re.search(r"入札[^0-9]{0,20}([\d,]+)", text)
    bids = bid_match.group(1) if bid_match else ""
    ship_match = re.search(r"送料[^0-9]{0,20}([\d,,]+)円", text)
    shipping = safe_int(ship_match.group(1)) if ship_match else 0
    seller_match = re.search(r'"seller":\{"id":"([^"]+)","rating":(\d+)', text)
    seller = seller_match.group(1) if seller_match else ""
    seller_score = seller_match.group(2) if seller_match else ""
    remaining = ""
    remain_match = re.search(r"(?:残り|終了まで)[^<]{0,20}(\d+日|\d+時間|\d+分)", text)
    if remain_match:
        remaining = remain_match.group(1)
    status = "出品中"
    if "このオークションは終了しています" in text:
        status = "已结束"
    item_type = detect_type(title)
    price = current_price or buyout_price or 0
    return {
        "platform": "Yahoo!オークション",
        "vol": query_vol,
        "status": status,
        "item_type": item_type,
        "title": title,
        "price_jpy": price,
        "buyout_jpy": buyout_price,
        "shipping_jpy": shipping,
        "landed_jpy": price + shipping,
        "price_cny": round((price + shipping) * JPY_TO_CNY, 1),
        "seller": seller,
        "seller_score": seller_score,
        "bids": bids,
        "remaining": remaining,
        "note": note_from_title(title),
        "risk_tags": "",
        "url": url,
        "keyword": keyword,
    }


def normalize_cached_row(row):
    title = row.get("title", "").strip()
    query_vol = row.get("query_vol", "")
    if not relevant_title(title, query_vol):
        return None
    item_type = detect_type(title)
    if item_type == "单卡":
        return None
    price = row.get("current_price") or row.get("buyout_price") or 0
    shipping = row.get("shipping_price") or 0
    return {
        "platform": "Yahoo!オークション",
        "vol": query_vol,
        "status": row.get("状态", "出品中"),
        "item_type": item_type,
        "title": title,
        "price_jpy": price,
        "buyout_jpy": row.get("buyout_price"),
        "shipping_jpy": shipping,
        "landed_jpy": price + shipping,
        "price_cny": round((price + shipping) * JPY_TO_CNY, 1),
        "seller": row.get("seller", ""),
        "seller_score": row.get("seller_rating", ""),
        "bids": row.get("bid_count", ""),
        "remaining": row.get("time_left", ""),
        "note": note_from_title(title),
        "risk_tags": "",
        "url": row.get("url", ""),
        "keyword": row.get("query", ""),
    }


def add_tags(rows):
    grouped = defaultdict(list)
    for row in rows:
        if row["item_type"] == "整盒" and row["price_jpy"]:
            grouped[row["vol"]].append(row["price_jpy"])
    medians = {k: median(v) for k, v in grouped.items() if v}
    for row in rows:
        tags = []
        if row["item_type"] not in ("整盒", "散包"):
            tags.append("搜索命中")
        base = medians.get(row["vol"])
        if base and row["price_jpy"] and row["price_jpy"] <= base * 0.6:
            tags.append("疑似低价")
        if row["bids"] and row["bids"] not in ("0", ""):
            tags.append("有竞价")
        row["risk_tags"] = " | ".join(tags)
        row["main_target"] = True


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(max(width + 2, 10), 42)


def header(ws, title, subtitle, cols):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A2"] = subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)


def write_sheet(ws, rows, headers, keys, title, subtitle):
    header(ws, title, subtitle, len(headers))
    fill = PatternFill("solid", fgColor="D9EAF7")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    r = 5
    for row in rows:
        for cidx, key in enumerate(keys, 1):
            val = row.get(key, "")
            if key in ("price_jpy", "buyout_jpy", "shipping_jpy", "landed_jpy"):
                val = fmt_jpy(val)
            elif key == "price_cny":
                val = fmt_cny(val)
            elif key == "url" and val:
                val = "🔗"
            ws.cell(row=r, column=cidx, value=val)
            if key == "url" and row.get("url"):
                ws.cell(row=r, column=cidx).hyperlink = row["url"]
                ws.cell(row=r, column=cidx).style = "Hyperlink"
        r += 1
    ws.freeze_panes = "A5"
    autosize(ws)


def build_summary(rows):
    out = []
    for vol in VOLS:
        items = [x for x in rows if x["vol"] == vol]
        best = min((x["landed_jpy"] for x in items if x["landed_jpy"]), default=None)
        out.append({
            "vol": vol,
            "count": len(items),
            "box_count": sum(1 for x in items if x["item_type"] == "整盒"),
            "pack_count": sum(1 for x in items if x["item_type"] == "散包"),
            "other_count": sum(1 for x in items if x["item_type"] not in ("整盒", "散包")),
            "best_jpy": best,
            "best_cny": round(best * JPY_TO_CNY, 1) if best else None,
        })
    return out


def build_workbook(rows):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    subtitle = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')} | 数据来源：Yahoo!オークション"
    ws = wb.create_sheet("总览")
    write_sheet(
        ws,
        build_summary(rows),
        ["分卷", "结果数", "整盒", "散包", "其他", "最低到手价JPY", "最低到手价CNY"],
        ["vol", "count", "box_count", "pack_count", "other_count", "best_jpy", "best_cny"],
        "监控｜雅虎拍卖宝石包｜总览",
        subtitle,
    )
    ws = wb.create_sheet("主清单")
    write_sheet(
        ws,
        rows,
        ["分卷", "状态", "类型", "标题", "当前价JPY", "一口价JPY", "运费JPY", "到手价JPY", "到手价CNY", "剩余时间", "出价数", "卖家", "卖家评价", "备注", "风险标记", "商品链接"],
        ["vol", "status", "item_type", "title", "price_jpy", "buyout_jpy", "shipping_jpy", "landed_jpy", "price_cny", "remaining", "bids", "seller", "seller_score", "note", "risk_tags", "url"],
        "监控｜雅虎拍卖宝石包｜主清单",
        subtitle,
    )
    ws = wb.create_sheet("全量结果")
    write_sheet(
        ws,
        rows,
        ["分卷", "状态", "类型", "标题", "当前价JPY", "一口价JPY", "运费JPY", "到手价JPY", "到手价CNY", "关键词", "商品链接"],
        ["vol", "status", "item_type", "title", "price_jpy", "buyout_jpy", "shipping_jpy", "landed_jpy", "price_cny", "keyword", "url"],
        "监控｜雅虎拍卖宝石包｜全量结果",
        subtitle,
    )
    ws = wb.create_sheet("杂质回收")
    write_sheet(
        ws,
        [],
        ["分卷", "状态", "类型", "标题", "商品链接"],
        ["vol", "status", "item_type", "title", "url"],
        "监控｜雅虎拍卖宝石包｜杂质回收",
        subtitle,
    )
    wb.save(OUT_FILE)


def main():
    rows = []
    cached_rows = load_cache_rows()
    if cached_rows:
        for row in cached_rows:
            item = normalize_cached_row(row)
            if item:
                rows.append(item)
    else:
        urls = search_urls()
        session = requests.Session()
        session.headers.update({"user-agent": "Mozilla/5.0"})
        for url, meta in urls.items():
            try:
                item = parse_detail(session, url, meta["query_vol"], meta["keyword"])
                if item:
                    rows.append(item)
            except Exception:
                continue
    rows = list(OrderedDict((row["url"], row) for row in rows if row.get("url")).values())
    add_tags(rows)
    build_workbook(rows)
    print(f"✅ 已保存：{OUT_FILE}")
    print(f"结果数：{len(rows)}")


if __name__ == "__main__":
    main()
