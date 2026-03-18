#!/usr/bin/env python3
import html
import os
import re
from collections import OrderedDict, defaultdict
from datetime import datetime
from statistics import median

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill


OUTPUT_DIR = os.path.expanduser("~/Desktop/销售总监報告")
OUT_FILE = os.path.join(OUTPUT_DIR, f"监控_速卖通宝石包_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
JPY_TO_CNY = 0.04334
VOLS = ["VOL.1", "VOL.2", "VOL.3", "VOL.4"]

# 扩展关键词库 - 覆盖用户要求的所有关键词
SEARCH_MATRIX = OrderedDict({
    "VOL.1": [
        "Pokemon Gem Pack VOL.1", "Pokemon Gem Pack VOL.1 box", 
        "Pokemon Gem Pack 1", "宝可梦 宝石包 第一弹",
        "ジェムパック vol.1", "ジェムパック 1",
        "gem pack pokemon vol 1", "gem pack pokemon 1",
        "CBB1C pokemon", "pokemon CBB1C"
    ],
    "VOL.2": [
        "Pokemon Gem Pack VOL.2", "Pokemon Gem Pack VOL.2 box",
        "Pokemon Gem Pack 2", "宝可梦 宝石包 第二弹",
        "ジェムパック vol.2", "ジェムパック 2",
        "gem pack pokemon vol 2", "gem pack pokemon 2",
        "CBB2C pokemon", "pokemon CBB2C"
    ],
    "VOL.3": [
        "Pokemon Gem Pack VOL.3", "Pokemon Gem Pack VOL.3 box",
        "Pokemon Gem Pack 3", "宝可梦 宝石包 第三弹",
        "ジェムパック vol.3", "ジェムパック 3",
        "gem pack pokemon vol 3", "gem pack pokemon 3",
        "CBB3C pokemon", "pokemon CBB3C"
    ],
    "VOL.4": [
        "Pokemon Gem Pack VOL.4", "Pokemon Gem Pack VOL.4 box",
        "Pokemon Gem Pack 4", "宝可梦 宝石包 第四弹",
        "ジェムパック vol.4", "ジェムパック 4",
        "gem pack pokemon vol 4", "gem pack pokemon 4",
        "CBB4C pokemon", "pokemon CBB4C"
    ],
})
MAX_KEYWORDS_PER_VOL = 4
TARGET_RESULTS_PER_VOL = 18

BOX_TERMS = ("box", "booster box", "ボックス", "箱", "sealed", "未開封", "シュリンク", "caixa", "caja", "1st edition", "first edition")
PACK_TERMS = ("pack", "booster pack", "パック", "1 pack", "10 pack", "12 pack", "16 pack", "18 pack", "pacote", "paquete", "sobre", "sobres")
SINGLE_RE = re.compile(r"\b(?:psa|bgs|ar|sar|sr)\b|シングル|ミラー", re.I)
BAD_TERMS = (
    "proxy", "diy", "badge", "album", "álbum", "storage", "book", "libro", "livro", "binder",
    "não original", "nao original", "non original", "unofficial", "no original",
    "handmade", "manual", "adesivo", "sticker", "custom", "emblema", "emblemas",
    "まとめ売り", "引退品", "GX", "VMAX", "Vstar", "TAG TEAM", "ジャンク", "訳あり"
)
SINGLE_BAD_RE = re.compile(
    r"(?:\bno\.?\s*\d+\b|\b\d+/\d+\b|#\s*\d+|full art|arte completo|carta individual|tarjeta individual|cart[aã]o individual)",
    re.I,
)
OXY_USER = "citibob_otC2c"
OXY_PASS = "oK1=Sp6n5n4iy5"
OXY_ENDPOINT = "https://realtime.oxylabs.io/v1/queries"


def fmt_jpy(v):
    return f"¥{int(v):,}" if isinstance(v, (int, float)) and v else ""


def fmt_cny(v):
    return f"¥{v:.1f}" if isinstance(v, (int, float)) and v is not None else ""


def detect_vol(title: str):
    t = title.upper()
    # 匹配 VOL.1, VOL1, VOL 1
    m = re.search(r"VOL[.\s]*([1-4])", t)
    if m:
        return f"VOL.{m.group(1)}"
    # 匹配 1st, 2nd, 3rd, 4th
    m = re.search(r"(\d)(?:ST|ND|RD|TH)\s*(?:PACK|BOX|GEM)", t)
    if m:
        return f"VOL.{m.group(1)}"
    # 匹配 CBB1C, CBB2C, CBB3C, CBB4C
    m = re.search(r"CBB([1-4])C", t)
    if m:
        return f"VOL.{m.group(1)}"
    # 匹配 GEM 1, GEM 2
    m = re.search(r"\bGEM\s*([1-4])\b", t)
    if m:
        return f"VOL.{m.group(1)}"
    # 匹配中文 第一弹, 第二弹等
    m = re.search(r"第([1-4])弹", t)
    if m:
        return f"VOL.{m.group(1)}"
    return ""


def detect_type(title: str):
    t = title.lower()
    if any(x in t for x in BOX_TERMS):
        return "整盒"
    if any(x in t for x in PACK_TERMS):
        return "散包"
    if SINGLE_RE.search(title):
        return "单卡"
    return "未分类"


def relevant_title(title: str, query_vol: str):
    t = title.lower()
    # 必须包含核心关键词
    core_terms = ("宝石包", "ジェム", "gem pack", "pokemon gem", "pokémon gem", "cbb", "gem", "pokemon")
    if not any(x in t for x in core_terms):
        return False
    # 过滤杂质
    if any(x in t for x in BAD_TERMS):
        return False
    if SINGLE_BAD_RE.search(title):
        return False
    # 必须包含盒或包关键词
    if not any(x in t for x in BOX_TERMS + PACK_TERMS):
        return False
    # 检测VOL并匹配查询
    detected = detect_vol(title)
    if detected and detected != query_vol:
        return False
    return True


def note_from_title(title: str):
    t = title.lower()
    notes = []
    m = re.search(r"(\d+)\s*(?:box|boxes|箱)", t)
    if m:
        notes.append(f"{m.group(1)}盒")
    m = re.search(r"(\d+)\s*(?:pack|packs|パック)", t)
    if m:
        notes.append(f"{m.group(1)}包")
    return " | ".join(notes)


def parse_price_to_jpy(raw: str):
    if not raw:
        return 0
    raw = raw.replace("\xa0", " ").strip()
    if "R$" in raw:
        nums = re.findall(r"[\d\.,]+", raw)
        if nums:
            val = float(nums[0].replace(".", "").replace(",", "."))
            return round(val * 29)
    if "$" in raw:
        nums = re.findall(r"[\d\.,]+", raw)
        if nums:
            val = float(nums[0].replace(",", ""))
            return round(val * 150)
    nums = re.findall(r"[\d\.,]+", raw)
    if nums:
        try:
            return int(float(nums[0].replace(",", "")))
        except Exception:
            return 0
    return 0


def fetch_search_html(query: str):
    payload = {
        "source": "aliexpress_search",
        "query": query,
        "user_agent_type": "desktop",
        "render": "html",
    }
    r = requests.post(OXY_ENDPOINT, auth=(OXY_USER, OXY_PASS), json=payload, timeout=30)
    r.raise_for_status()
    return r.json()["results"][0]["content"]


def parse_cards(search_html: str, query_vol: str, keyword: str):
    rows = []
    for m in re.finditer(r'<a class="[^"]*search-card-item[^"]*" href="([^"]+)"', search_html):
        href = html.unescape(m.group(1))
        if href.startswith("//"):
            href = "https:" + href
        start = m.start()
        end = search_html.find("</a>", start)
        if end == -1:
            continue
        block = search_html[start:end]
        title_match = re.search(r'<h3 class="[^"]*">(.+?)</h3>', block, re.S)
        title = html.unescape(re.sub(r"<.*?>", "", title_match.group(1))).strip() if title_match else ""
        if not title or not relevant_title(title, query_vol):
            continue
        price_match = re.search(r'aria-label="([^"]+)"', block)
        price_text = html.unescape(price_match.group(1)).strip() if price_match else ""
        sold_match = re.search(r'(\d+\s*(?:sold|vendidos|sales))', block, re.I)
        sold = sold_match.group(1) if sold_match else ""
        item_type = detect_type(title)
        price_jpy = parse_price_to_jpy(price_text)
        rows.append({
            "vol": query_vol,
            "status": "販売中",
            "item_type": item_type,
            "title": title,
            "price_jpy": price_jpy,
            "shipping_jpy": 0,
            "landed_jpy": price_jpy,
            "price_cny": round(price_jpy * JPY_TO_CNY, 1),
            "seller": "",
            "sales": sold,
            "note": note_from_title(title),
            "risk_tags": "",
            "url": href.split("?")[0],
            "keyword": keyword,
        })
    return rows


def add_tags(rows):
    grouped = defaultdict(list)
    for row in rows:
        if row["item_type"] == "整盒" and row["price_jpy"]:
            grouped[row["vol"]].append(row["price_jpy"])
    medians = {k: median(v) for k, v in grouped.items() if v}
    for row in rows:
        tags = []
        if row["item_type"] not in ("整盒", "散包"):
            tags.append("搜索命中")
        base = medians.get(row["vol"])
        if base and row["price_jpy"] and row["price_jpy"] <= base * 0.6:
            tags.append("疑似低价")
        row["risk_tags"] = " | ".join(tags)


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column_letter] = max(widths.get(cell.column_letter, 0), len(str(cell.value)))
    for col, width in widths.items():
        ws.column_dimensions[col].width = min(max(width + 2, 10), 42)


def header(ws, title, subtitle, cols):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws["A2"] = subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)


def write_sheet(ws, rows, headers, keys, title, subtitle):
    header(ws, title, subtitle, len(headers))
    fill = PatternFill("solid", fgColor="D9EAF7")
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    r = 5
    for row in rows:
        for cidx, key in enumerate(keys, 1):
            val = row.get(key, "")
            if key in ("price_jpy", "shipping_jpy", "landed_jpy"):
                val = fmt_jpy(val)
            elif key == "price_cny":
                val = fmt_cny(val)
            elif key == "url" and val:
                val = "链接"
            ws.cell(row=r, column=cidx, value=val)
            if key == "url" and row.get("url"):
                ws.cell(row=r, column=cidx).hyperlink = row["url"]
                ws.cell(row=r, column=cidx).style = "Hyperlink"
        r += 1
    ws.freeze_panes = "A5"
    autosize(ws)


def build_summary(rows):
    out = []
    for vol in VOLS:
        items = [x for x in rows if x["vol"] == vol]
        best = min((x["landed_jpy"] for x in items if x["landed_jpy"]), default=None)
        out.append({
            "vol": vol,
            "count": len(items),
            "box_count": sum(1 for x in items if x["item_type"] == "整盒"),
            "pack_count": sum(1 for x in items if x["item_type"] == "散包"),
            "other_count": sum(1 for x in items if x["item_type"] not in ("整盒", "散包")),
            "best_jpy": best,
            "best_cny": round(best * JPY_TO_CNY, 1) if best else None,
        })
    return out


def build_workbook(rows):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    subtitle = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')} | 数据来源：AliExpress + Oxylabs"
    ws = wb.create_sheet("总览")
    write_sheet(
        ws,
        build_summary(rows),
        ["分卷", "结果数", "整盒", "散包", "其他", "最低到手价JPY", "最低到手价CNY"],
        ["vol", "count", "box_count", "pack_count", "other_count", "best_jpy", "best_cny"],
        "监控｜速卖通宝石包｜总览",
        subtitle,
    )
    ws = wb.create_sheet("主清单")
    write_sheet(
        ws,
        rows,
        ["分卷", "状态", "类型", "标题", "商品价JPY", "运费JPY", "到手价JPY", "到手价CNY", "销量", "备注", "风险标记", "商品链接"],
        ["vol", "status", "item_type", "title", "price_jpy", "shipping_jpy", "landed_jpy", "price_cny", "sales", "note", "risk_tags", "url"],
        "监控｜速卖通宝石包｜主清单",
        subtitle,
    )
    ws = wb.create_sheet("全量结果")
    write_sheet(
        ws,
        rows,
        ["分卷", "状态", "类型", "标题", "商品价JPY", "到手价JPY", "到手价CNY", "关键词", "商品链接"],
        ["vol", "status", "item_type", "title", "price_jpy", "landed_jpy", "price_cny", "keyword", "url"],
        "监控｜速卖通宝石包｜全量结果",
        subtitle,
    )
    ws = wb.create_sheet("杂质回收")
    write_sheet(
        ws,
        [],
        ["分卷", "状态", "类型", "标题", "商品链接"],
        ["vol", "status", "item_type", "title", "url"],
        "监控｜速卖通宝石包｜杂质回收",
        subtitle,
    )
    wb.save(OUT_FILE)


def main():
    rows = []
    errors = []
    for vol, keywords in SEARCH_MATRIX.items():
        vol_rows = []
        for keyword in keywords[:MAX_KEYWORDS_PER_VOL]:
            existing = list(OrderedDict((row["url"], row) for row in vol_rows if row.get("url")).values())
            if len(existing) >= TARGET_RESULTS_PER_VOL:
                print(f"✓ {vol}: 已达到 {TARGET_RESULTS_PER_VOL} 条，停止继续扩搜")
                break
            try:
                html_content = fetch_search_html(keyword)
                parsed = parse_cards(html_content, vol, keyword)
                vol_rows.extend(parsed)
                print(f"✓ {vol} - {keyword}: {len(parsed)} items")
            except Exception as e:
                errors.append(f"{vol}/{keyword}: {str(e)}")
                print(f"✗ {vol} - {keyword}: {str(e)}")
                continue
        rows.extend(list(OrderedDict((row["url"], row) for row in vol_rows if row.get("url")).values()))

    # 去重
    rows = list(OrderedDict((row["url"], row) for row in rows if row.get("url")).values())
    
    if len(rows) < 3:
        raise RuntimeError(f"速卖通有效结果不足({len(rows)})，停止输出空报告")
    
    add_tags(rows)
    build_workbook(rows)
    print(f"\n✅ 已保存：{OUT_FILE}")
    print(f"结果数：{len(rows)}")
    if errors:
        print(f"错误：{errors}")


if __name__ == "__main__":
    main()
