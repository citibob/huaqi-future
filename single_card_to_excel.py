#!/usr/bin/env python3
"""单卡监控 - MD转Excel"""
import os
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook

OUTPUT_DIR = "/Users/hm/Desktop/销售总监報告"

def parse_md(md_path):
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()
    
    sections = {}
    current_section = None
    current_items = []
    
    for line in content.split("\n"):
        if line.startswith("## "):
            if current_section:
                sections[current_section] = current_items
            current_section = line.replace("## ", "").strip()
            current_items = []
        elif line.strip().startswith("1. ") or line.strip().startswith("2. ") or line.strip().startswith("3. "):
            match = re.match(r"\d+\.\s+¥(\d+)[,，]\s*~¥(\d+)[,，]\s+-\s+(.+)", line.strip())
            if match:
                jpy = int(match.group(1))
                cny = float(match.group(2))
                title = match.group(3)
                current_items.append({"title": title, "jpy": jpy, "cny": cny, "url": ""})
        elif line.strip().startswith("http"):
            if current_items:
                current_items[-1]["url"] = line.strip()
    
    if current_section:
        sections[current_section] = current_items
    
    return sections

def generate_excel(sections, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "单卡监控"
    
    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = "🃏   单卡价格监控报告  ·  PSA10 AR/SAR"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # 副标题
    ws.merge_cells("A2:F2")
    ws["A2"] = f"  生成时间：{datetime.now().strftime('%Y-%m-%d')}  |  汇率：1 JPY = 0.0434 CNY"
    ws["A2"].font = Font(name="Arial", size=10)
    
    current_row = 4
    
    # 样式
    section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    section_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    col_header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    
    for section_name, items in sections.items():
        # 区块标题
        ws.merge_cells(f"A{current_row}:F{current_row}")
        cell = ws[f"A{current_row}"]
        cell.value = f"   🃏  {section_name}"
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        
        # 列标题
        col_headers = ["#", "商品标题", "日元价格", "人民币(≈)", "商品链接", "备注"]
        for col, h in enumerate(col_headers, 1):
            cell = ws.cell(current_row, col)
            cell.value = h
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        
        # 数据
        for idx, item in enumerate(items[:10], 1):
            ws.cell(current_row, 1, str(idx))
            ws.cell(current_row, 2, item.get("title", "")[:50])
            ws.cell(current_row, 3, item.get("jpy", 0))
            ws.cell(current_row, 4, item.get("cny", 0))
            ws.cell(current_row, 5, "🔗 查看")
            ws.cell(current_row, 6, "—")
            current_row += 1
        
        current_row += 1
    
    # 设置列宽
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    
    wb.save(output_path)

def main():
    today = datetime.now().strftime("%Y-%m-%d")
    md_path = f"{OUTPUT_DIR}/单卡监控_{today}.md"
    xlsx_path = f"{OUTPUT_DIR}/单卡报告_{today}.xlsx"
    
    if not os.path.exists(md_path):
        print(f"❌ 找不到: {md_path}")
        return
    
    sections = parse_md(md_path)
    generate_excel(sections, xlsx_path)
    print(f"✅ 已保存: {xlsx_path}")
    print(f"共 {len(sections)} 个区块")

if __name__ == "__main__":
    main()
